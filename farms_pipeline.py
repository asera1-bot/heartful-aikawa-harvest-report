#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# Constants / Regex
# ============================================================
KI_LABEL_RE = re.compile(r"^\s*(\d+)\s*期\s*$")


# ============================================================
# 期判定（10/1〜翌9/30）
# ============================================================
def calc_ki(d) -> int:
    """
    期判定（10/1〜翌9/30）
    例）2024-10-01〜2025-09-30 -> 58期
       2025-10-01〜2026-09-30 -> 59期
    """
    ts = pd.to_datetime(d, errors="coerce")
    if pd.isna(ts):
        return -1
    ki_year = ts.year + 1 if ts.month >= 10 else ts.year  # 終了年
    return ki_year - 1967  # ki_year=2025 -> 58期


# ============================================================
# Utils
# ============================================================
def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def resolve_path(path_str: str, label: str = "file") -> Path:
    p = Path(path_str).expanduser()
    if not p.exists():
        eprint(f"[ERROR] {label} not found: {p}")
        eprint(f"[INFO] cwd: {Path.cwd()}")
        sys.exit(2)
    return p.resolve()


def ensure_outdir(path_str: str) -> Path:
    outdir = Path(path_str).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)
    return outdir


def parse_target_months(month: Optional[int], months: Optional[str]) -> List[int]:
    if month is not None:
        if not (1 <= month <= 12):
            raise SystemExit(f"[ERROR] invalid --month: {month} (1-12)")
        return [month]

    s = (months or "").strip().lower()
    if s == "all":
        return list(range(10, 13)) + list(range(1, 10))

    items = [x.strip() for x in s.split(",") if x.strip()]
    if not items:
        raise SystemExit("[ERROR] invalid --months (use all or like 10,11,12)")

    out: List[int] = []
    for x in items:
        m = int(x)
        if not (1 <= m <= 12):
            raise SystemExit(f"[ERROR] invalid month in --months: {m} (1-12)")
        out.append(m)
    return out


def parse_kis(kis_str: str | None, fallback_ki: int) -> List[int]:
    if kis_str is None or str(kis_str).strip() == "":
        return [int(fallback_ki)]
    items = [x.strip() for x in str(kis_str).split(",") if x.strip()]
    out: List[int] = []
    for x in items:
        m = re.search(r"\d+", x)
        if not m:
            raise SystemExit(f"[ERROR] invalid --kis item: {x}")
        out.append(int(m.group()))
    # 重複除去・昇順
    return sorted(set(out))


def make_suffix(target_months: List[int]) -> str:
    if len(target_months) == 12 and target_months == (list(range(10, 13)) + list(range(1, 10))):
        return "all"
    if len(target_months) == 1:
        return f"{target_months[0]:02d}"
    return "m" + "-".join(f"{m:02d}" for m in target_months)


def make_report_outname(ki: int, target_months: List[int]) -> str:
    return f"report_{ki}ki_{make_suffix(target_months)}.xlsx"


def make_unmapped_outname(ki: int, target_months: List[int]) -> str:
    return f"unmapped_{ki}ki_{make_suffix(target_months)}.xlsx"


def norm_key(s: str) -> str:
    """
    テンプレの見た目ゆれ吸収用（スペース/改行/全角→半角/スラッシュ等）
    """
    s = unicodedata.normalize("NFKC", str(s or ""))
    s = s.replace("／", "/")
    s = s.replace("\u3000", " ")
    s = s.replace("\n", "")
    s = re.sub(r"\s+", "", s)  # 空白は全削除
    return s.strip()


# ============================================================
# IO dataclasses
# ============================================================
@dataclass
class Inputs:
    norm_src: Path
    master: Path
    harvest_template: Path
    report_template: Path


# ============================================================
# Master
# ============================================================
def validate_master(master_path: Path) -> None:
    xls = pd.ExcelFile(master_path)
    needed = {"企業名", "収穫野菜名"}
    got = set(xls.sheet_names)
    missing = needed - got
    if missing:
        raise SystemExit(f"[ERROR] master missing sheets: {sorted(missing)}")


def read_master_maps(master_path: Path) -> tuple[Dict[str, str], Dict[str, Dict[str, str]]]:
    # 企業名: A=元, B=変換後（×は除外）
    df_c = pd.read_excel(master_path, sheet_name="企業名", usecols=[0, 1], dtype=str).fillna("")
    company_map = {a.strip(): b.strip() for a, b in zip(df_c.iloc[:, 0], df_c.iloc[:, 1]) if a.strip()}

    # 収穫野菜名: A=元, B=suffix, C=bucket
    df_v = pd.read_excel(master_path, sheet_name="収穫野菜名", usecols=[0, 1, 2], dtype=str).fillna("")
    veg_map: Dict[str, Dict[str, str]] = {}
    for a, b, c in zip(df_v.iloc[:, 0], df_v.iloc[:, 1], df_v.iloc[:, 2]):
        k = a.strip()
        if not k:
            continue
        veg_map[k] = {"suffix": b.strip(), "bucket": c.strip()}
    return company_map, veg_map


# ============================================================
# Load norm
# ============================================================
def load_norm(norm_src: Path) -> pd.DataFrame:
    df = pd.read_excel(norm_src)

    def norm_col(s):
        s = unicodedata.normalize("NFKC", str(s)).strip()
        s = re.sub(r"\s+", "", s)
        return s

    df = df.rename(columns={c: norm_col(c) for c in df.columns})

    if "収穫量(g)" not in df.columns:
        raise SystemExit(f"[ERROR] 収穫量(g) が見つかりません。cols={list(df.columns)}")

    df["収穫量(g)"] = pd.to_numeric(df["収穫量(g)"], errors="coerce")
    return df


# ============================================================
# Normalize / Map
# ============================================================
def normalize_and_map(df_norm: pd.DataFrame, master_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    required_cols = ["収穫日", "得意先名", "収穫物", "収穫量(g)"]
    missing = [c for c in required_cols if c not in df_norm.columns]
    if missing:
        raise SystemExit(f"[ERROR] 正規化.xlsx に必要列がありません: {missing}")

    company_map, veg_map = read_master_maps(master_path)

    rows_ok: list[dict] = []
    rows_ng: list[dict] = []

    for _, r in df_norm.iterrows():
        harvest_date = r["収穫日"]
        company_raw = str(r["得意先名"]).strip()
        crop_raw = str(r["収穫物"]).strip()
        amount = r["収穫量(g)"]

        # 企業名変換
        if company_raw not in company_map:
            rows_ng.append(
                {"type": "company_unmapped", "得意先名": company_raw, "収穫物": crop_raw, "収穫日": harvest_date, "収穫量(g)": amount}
            )
            continue

        company_conv = company_map[company_raw]
        if company_conv == "×":
            rows_ng.append(
                {"type": "company_excluded(×)", "得意先名": company_raw, "変換後名": company_conv, "収穫物": crop_raw, "収穫日": harvest_date, "収穫量(g)": amount}
            )
            continue

        # 収穫物変換（suffix/bucket）
        if crop_raw not in veg_map:
            rows_ng.append(
                {"type": "crop_unmapped", "得意先名": company_raw, "変換後名": company_conv, "収穫物": crop_raw, "収穫日": harvest_date, "収穫量(g)": amount}
            )
            continue

        suffix = veg_map[crop_raw]["suffix"]
        bucket = veg_map[crop_raw]["bucket"]

        # 強制救済（マスタが揺れても加工BLは加工用に寄せる）
        if "加工BL" in norm_key(crop_raw):
            bucket = "加工用"
            suffix = "/加工BL"

        if not bucket:
            rows_ng.append(
                {"type": "bucket_unmapped", "得意先名": company_raw, "変換後名": company_conv, "収穫物": crop_raw, "suffix": suffix, "収穫日": harvest_date, "収穫量(g)": amount}
            )
            continue

        # company_proc（確定ロジック）
        if bucket == "べビーリーフ":
            company_proc = company_conv
        else:
            company_proc = f"{company_conv}{suffix}"

        rows_ok.append(
            {
                "収穫日": harvest_date,
                "company_raw": company_raw,
                "company_conv": company_conv,
                "crop_raw": crop_raw,
                "suffix": suffix,
                "bucket": bucket,
                "company_proc": company_proc,
                "収穫量(g)": amount,
            }
        )

    df_ok = pd.DataFrame(rows_ok)
    df_ng = pd.DataFrame(rows_ng)

    # 収穫量を数値化（壊れた値はunmappedへ）
    if not df_ok.empty:
        df_ok["収穫量(g)"] = pd.to_numeric(df_ok["収穫量(g)"], errors="coerce")
        bad = df_ok["収穫量(g)"].isna()
        if bad.any():
            moved = df_ok[bad].copy()
            moved["type"] = "amount_not_number"
            df_ng = pd.concat([df_ng, moved], ignore_index=True)
            df_ok = df_ok[~bad].copy()

    return df_ok, df_ng


# ============================================================
# Unmapped output
# ============================================================
def write_unmapped(path: Path, df_unmapped: pd.DataFrame) -> None:
    if df_unmapped is None or df_unmapped.empty:
        pd.DataFrame([{"status": "no unmapped rows"}]).to_excel(path, index=False)
        return
    df_unmapped.to_excel(path, index=False)


# ============================================================
# Harvest_all (UI: 月別シートへ羅列)
# ============================================================
def clear_harvest_rows(ws, start_row: int = 2, col_start: int = 1, col_end: int = 6) -> None:
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).value = None


def write_harvest_all_by_month_sheets(
    template_path: Path,
    out_path: Path,
    df_ok: pd.DataFrame,
    target_kis: List[int],
) -> None:
    wb = load_workbook(template_path)

    df_m = df_ok.copy()
    df_m["収穫日"] = pd.to_datetime(df_m["収穫日"], errors="coerce")
    df_m = df_m.dropna(subset=["収穫日"]).copy()

    # ki列を作って target_kis で絞る（年度混入を防止）
    df_m["ki"] = df_m["収穫日"].apply(calc_ki)
    df_m = df_m[df_m["ki"].isin(list(map(int, target_kis)))].copy()

    sheet_months = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    start_row = 2

    for m in sheet_months:
        sheet_name = f"{m}月"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        clear_harvest_rows(ws, start_row=start_row, col_start=1, col_end=6)

        dmm = df_m[df_m["収穫日"].dt.month == m].copy()
        if dmm.empty:
            continue

        dmm = dmm.sort_values(["収穫日", "company_proc", "crop_raw"], kind="mergesort")

        r = start_row
        harvest_id = 1
        for _, row in dmm.iterrows():
            ws.cell(row=r, column=1).value = harvest_id
            ws.cell(row=r, column=2).value = "愛川"
            ws.cell(row=r, column=3).value = row["収穫日"].date()
            ws.cell(row=r, column=4).value = row.get("company_proc", "")
            ws.cell(row=r, column=5).value = row.get("crop_raw", "")
            ws.cell(row=r, column=6).value = float(row["収穫量(g)"]) if pd.notna(row["収穫量(g)"]) else None
            r += 1
            harvest_id += 1

    wb.save(out_path)


# ============================================================
# Report helpers
# ============================================================
def pick_sheet_name(wb, keyword: str) -> str:
    for n in wb.sheetnames:
        if keyword in n:
            return n
    raise SystemExit(f"[ERROR] report sheet not found by keyword: {keyword} (sheets={wb.sheetnames})")


def find_month_columns(ws) -> Dict[int, int]:
    labels = {
        10: "10月", 11: "11月", 12: "12月",
        1: "1月", 2: "2月", 3: "3月",
        4: "4月", 5: "5月", 6: "6月",
        7: "7月", 8: "8月", 9: "9月",
    }

    for r in range(1, 201):
        row_vals = []
        for c in range(1, 51):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                v = v.strip()
            row_vals.append(v)

        if "10月" in row_vals and "9月" in row_vals:
            month_to_col: Dict[int, int] = {}
            for m, lab in labels.items():
                if lab in row_vals:
                    month_to_col[m] = row_vals.index(lab) + 1
            return month_to_col

    raise SystemExit("[ERROR] report template: month header row (10月..9月) not found")


def _is_total_row_name(name: str) -> bool:
    s = (name or "").strip()
    return ("企業総合計" in s) or ("月間総合計" in s) or ("総合計" in s) or ("合計" in s)


def safe_set_value(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = None
    cell.value = value


def recalc_totals(ws, row_idx: int) -> None:
    """
    report_template の合計列を “値” で埋める。
    - 10月〜3月: 4〜9列 → 上期(10列)
    - 4月〜9月: 11〜16列 → 下期(17列)
    - 合計: 18列
    """
    first_half_cols = range(4, 10)    # 4..9
    second_half_cols = range(11, 17)  # 11..16

    first_half = 0.0
    second_half = 0.0

    for col in first_half_cols:
        v = ws.cell(row=row_idx, column=col).value
        if isinstance(v, (int, float)):
            first_half += float(v)

    for col in second_half_cols:
        v = ws.cell(row=row_idx, column=col).value
        if isinstance(v, (int, float)):
            second_half += float(v)

    ws.cell(row=row_idx, column=10).value = round(first_half, 2)
    ws.cell(row=row_idx, column=17).value = round(second_half, 2)
    ws.cell(row=row_idx, column=18).value = round(first_half + second_half, 2)


# ============================================================
# merged cell (高速化: キャッシュ)
# ============================================================
def _build_merged_cache(ws):
    # openpyxl range list
    ranges = list(ws.merged_cells.ranges)
    cache: Dict[Tuple[int, int], Tuple[int, int]] = {}  # (r,c) -> (topr, topc)
    return ranges, cache


def _merged_topleft_value(ws, r: int, c: int, merged_ranges, merged_cache) -> object:
    """
    結合セル対策（キャッシュ付き）
    """
    key = (r, c)
    if key in merged_cache:
        tr, tc = merged_cache[key]
        return ws.cell(tr, tc).value

    coord = ws.cell(r, c).coordinate
    for rng in merged_ranges:
        if coord in rng:
            merged_cache[key] = (rng.min_row, rng.min_col)
            return ws.cell(rng.min_row, rng.min_col).value

    merged_cache[key] = (r, c)
    return ws.cell(r, c).value


# ============================================================
# ★ 期ブロック検出（完成版：スコアリングで本物を選ぶ）
# ============================================================
def _find_all_ki_label_rows(ws, merged_ranges, merged_cache, max_row: int) -> List[Tuple[int, int]]:
    ki_rows: List[Tuple[int, int]] = []
    for r in range(1, max_row + 1):
        v = _merged_topleft_value(ws, r, 3, merged_ranges, merged_cache)  # C列
        if v is None:
            continue
        s = v.strip() if isinstance(v, str) else ""
        m = KI_LABEL_RE.match(s)
        if not m:
            continue
        ki_rows.append((r, int(m.group(1))))
    ki_rows.sort(key=lambda x: x[0])
    return ki_rows


def _looks_like_company_row(ws, merged_ranges, merged_cache, r: int) -> bool:
    bad_words = ("増減", "稼働", "KPI", "目標", "実績", "差", "前年差", "前月差", "前年差異")

    # B列優先、無ければA列
    for c in (2, 1):
        v = _merged_topleft_value(ws, r, c, merged_ranges, merged_cache)
        if isinstance(v, str):
            name = v.strip()
            if not name:
                continue
            if any(w in name for w in bad_words):
                return False
            if _is_total_row_name(name):
                return False
            if KI_LABEL_RE.match(name):
                return False
            return True
    return False


def _choose_best_label_row_for_ki(
    ws,
    merged_ranges,
    merged_cache,
    ki_rows: List[Tuple[int, int]],
    target_ki: int,
    max_row: int,
    debug: bool = False,
) -> Optional[int]:
    all_label_rows = [r for r, _ in ki_rows]
    candidates = [r for r, k in ki_rows if k == target_ki]

    best_row = None
    best_score = -1

    # “本物ブロック”は短すぎない前提（ここが誤判定の核心）
    MIN_SPAN = 50
    WINDOW = 20

    for r in candidates:
        nxt = None
        for rr in all_label_rows:
            if rr > r:
                nxt = rr
                break
        end = (nxt - 1) if nxt else max_row
        span = end - r

        if span < MIN_SPAN:
            continue

        hit = 0
        for i in range(1, WINDOW + 1):
            if r + i > max_row:
                break
            if _looks_like_company_row(ws, merged_ranges, merged_cache, r + i):
                hit += 1

        # スコア：企業行の密度を最優先、次にスパン
        score = hit * 1000 + span

        if debug:
            print(f"[DEBUG] ki={target_ki} label_row={r} span={span} company_hit={hit} score={score}")

        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def build_ki_block_ranges(
    ws,
    target_kis: List[int],
    max_row: int = 4000,
    debug: bool = False,
) -> Dict[int, Tuple[int, int]]:
    merged_ranges, merged_cache = _build_merged_cache(ws)

    ki_rows = _find_all_ki_label_rows(ws, merged_ranges, merged_cache, max_row=max_row)
    if not ki_rows:
        raise SystemExit("[ERROR] cannot find any ki label rows in template (C column).")

    chosen_label: Dict[int, int] = {}
    for ki in target_kis:
        rr = _choose_best_label_row_for_ki(ws, merged_ranges, merged_cache, ki_rows, ki, max_row=max_row, debug=debug)
        if rr is not None:
            chosen_label[int(ki)] = rr

    if debug:
        print("[DEBUG] chosen_label_rows:", chosen_label)

    if not chosen_label:
        raise SystemExit("[ERROR] cannot choose ki label rows (all candidates too weak).")

    items = sorted(chosen_label.items(), key=lambda x: x[1])  # (ki, row)
    ranges: Dict[int, Tuple[int, int]] = {}
    for i, (ki, row) in enumerate(items):
        start = row + 1
        end = (items[i + 1][1] - 1) if i + 1 < len(items) else max_row
        ranges[int(ki)] = (start, end)

    if debug:
        print("[DEBUG] ki block ranges:", ranges)

    return ranges


def build_company_rows_by_ki_block(
    ws,
    target_kis: List[int],
    max_row: int = 4000,
    debug: bool = False,
) -> tuple[Dict[int, Dict[str, int]], Dict[int, int]]:
    """
    期ブロック（スコアリングで選んだ58/59）を対象に、企業行を抽出する。
    戻り値:
      - row_map_by_ki[ki][norm_company] = row_idx
      - total_row_by_ki[ki] = 合計行 row_idx（見つかった場合）
    """
    merged_ranges, merged_cache = _build_merged_cache(ws)

    block_ranges = build_ki_block_ranges(ws, target_kis=target_kis, max_row=max_row, debug=debug)

    row_map_by_ki: Dict[int, Dict[str, int]] = {int(k): {} for k in target_kis}
    total_row_by_ki: Dict[int, int] = {}

    def get_name(r: int) -> str:
        for cc in (2, 1):
            v = _merged_topleft_value(ws, r, cc, merged_ranges, merged_cache)
            if isinstance(v, str) and v.strip():
                return v.strip()
        return ""

    bad_words = ("増減", "稼働", "KPI", "目標", "実績", "差", "前年差", "前月差", "前年差異")

    for ki, (start, end) in block_ranges.items():
        if int(ki) not in row_map_by_ki:
            continue
        for r in range(start, min(end, max_row) + 1):
            name = get_name(r)
            if not name:
                continue
            if any(w in name for w in bad_words):
                continue

            if _is_total_row_name(name):
                total_row_by_ki.setdefault(int(ki), r)
                continue

            key = norm_key(name)
            row_map_by_ki[int(ki)].setdefault(key, r)

    if debug:
        for ki in sorted(row_map_by_ki.keys()):
            print(f"[DEBUG] ki={ki} company_rows={len(row_map_by_ki[ki])} total_row={total_row_by_ki.get(ki)}")

    return row_map_by_ki, total_row_by_ki


def clear_report_cells_for_rows(
    ws,
    month_to_col: Dict[int, int],
    rows: List[int],
    target_months: List[int],
) -> None:
    for r in rows:
        for m in target_months:
            col = month_to_col.get(int(m))
            if col:
                safe_set_value(ws, r, col, None)
        safe_set_value(ws, r, 10, None)
        safe_set_value(ws, r, 17, None)
        safe_set_value(ws, r, 18, None)


def recalc_total_rows_by_blocks(
    ws,
    month_to_col: Dict[int, int],
    row_map_by_ki: Dict[int, Dict[str, int]],
    total_row_by_ki: Dict[int, int],
) -> None:
    for ki, total_row in total_row_by_ki.items():
        rows = list(row_map_by_ki.get(int(ki), {}).values())
        if not rows:
            continue

        for m, col in month_to_col.items():
            s = 0.0
            for r in rows:
                v = ws.cell(r, col).value
                if isinstance(v, (int, float)):
                    s += float(v)
            safe_set_value(ws, total_row, col, round(s, 2))

        recalc_totals(ws, total_row)


# ============================================================
# Report writer
# ============================================================
def write_report_template(
    report_template: Path,
    out_path: Path,
    df_ok: pd.DataFrame,
    target_months: List[int],
    target_kis: List[int],
    report_unit: str,
    debug: bool,
) -> pd.DataFrame:
    """
    report_template を読み込み、df_ok を元に「対象期×対象月」だけ値上書きする。
    UI崩壊防止：書式/罫線/結合は触らず value だけ上書きする
    戻り値: テンプレ未登録/マスタ未登録などの unmapped 行（DataFrame）
    """
    wb = load_workbook(report_template)

    sheet_by_bucket = {
        "べビーリーフ": pick_sheet_name(wb, "企業月間生産量（ベビーリーフ）"),
        "べビーリーフ以外": pick_sheet_name(wb, "企業月間生産量（ベビーリーフ以外）"),
        "加工用": pick_sheet_name(wb, "加工用"),
    }

    df = df_ok.copy()
    df["収穫日"] = pd.to_datetime(df["収穫日"], errors="coerce")
    df = df.dropna(subset=["収穫日"]).copy()
    df["month"] = df["収穫日"].dt.month
    df["ki"] = df["収穫日"].apply(calc_ki)

    df = df[df["month"].isin(list(map(int, target_months)))].copy()
    df = df[df["ki"].isin(list(map(int, target_kis)))].copy()

    g = (
        df.groupby(["bucket", "company_proc", "ki", "month"], as_index=False)["収穫量(g)"]
        .sum()
        .rename(columns={"収穫量(g)": "amount_g"})
    )

    if report_unit == "kg":
        g["amount"] = (g["amount_g"].astype(float) / 1000.0).round(2)
    else:
        g["amount"] = g["amount_g"].astype(float)

    unmapped_rows: list[dict] = []

    for bucket, sheet_name in sheet_by_bucket.items():
        ws = wb[sheet_name]
        month_to_col = find_month_columns(ws)

        gb = g[g["bucket"] == bucket]
        if gb.empty:
            if debug:
                print(f"[DEBUG] bucket={bucket} gb_rows=0 (skip)")
            continue

        kis_to_write = sorted(set(int(x) for x in gb["ki"].unique()))
        months_to_write = sorted(set(int(x) for x in gb["month"].unique()))

        row_map_by_ki, total_row_by_ki = build_company_rows_by_ki_block(
            ws,
            target_kis=target_kis,
            max_row=4000,
            debug=debug,
        )

        # クリア対象を「書ける行」だけに限定
        rows_to_clear: list[int] = []
        for ki2 in kis_to_write:
            mki = row_map_by_ki.get(int(ki2), {})
            for company in gb["company_proc"].unique():
                key = norm_key(str(company).strip())
                r = mki.get(key)
                if r:
                    rows_to_clear.append(r)
        rows_to_clear = sorted(set(rows_to_clear))

        clear_report_cells_for_rows(
            ws,
            month_to_col=month_to_col,
            rows=rows_to_clear,
            target_months=months_to_write,
        )

        touched_rows: set[int] = set()

        for _, rr in gb.iterrows():
            company = str(rr["company_proc"]).strip()
            ki2 = int(rr["ki"])
            m = int(rr["month"])
            amt = float(rr["amount"])

            if not company or ki2 <= 0:
                continue

            if m not in month_to_col:
                unmapped_rows.append(
                    {"type": "template_unmapped(month_col)", "bucket": bucket, "company_proc": company, "ki": ki2, "month": m, "amount": amt}
                )
                continue

            key = norm_key(company)
            row = row_map_by_ki.get(ki2, {}).get(key)
            if not row:
                unmapped_rows.append(
                    {"type": "template_unmapped(company_row)", "bucket": bucket, "company_proc": company, "ki": ki2, "month": m, "amount": amt}
                )
                continue

            col = month_to_col[m]
            safe_set_value(ws, row, col, amt)
            touched_rows.add(row)

        for row_idx in touched_rows:
            recalc_totals(ws, row_idx)

        recalc_total_rows_by_blocks(
            ws,
            month_to_col=month_to_col,
            row_map_by_ki=row_map_by_ki,
            total_row_by_ki=total_row_by_ki,
        )

        if debug:
            print(
                f"[DEBUG] bucket={bucket} gb_rows={len(gb)} "
                f"rows_to_clear={len(rows_to_clear)} touched_rows={len(touched_rows)} "
                f"unmapped_total={len(unmapped_rows)}"
            )

    wb.save(out_path)
    return pd.DataFrame(unmapped_rows)


# ============================================================
# Pipeline
# ============================================================
def run_pipeline(
    *,
    inputs: Inputs,
    outdir: Path,
    ki: int,
    target_months: List[int],
    target_kis: List[int],
    do_validate: bool,
    report_unit: str,
    debug: bool,
) -> None:
    if do_validate:
        validate_master(inputs.master)

    report_out = outdir / make_report_outname(ki, target_months)
    unmapped_out = outdir / make_unmapped_outname(ki, target_months)

    df = load_norm(inputs.norm_src)
    df_ok, df_ng = normalize_and_map(df, inputs.master)

    if df_ok is None or df_ok.empty:
        raise SystemExit("[ERROR] df_ok is empty. check master mapping or norm data.")

    # harvest_all（監査用：UI提出形式）※年度混入を防ぐため、target_kis に [ki] を渡す
    all_path = outdir / f"harvest_{ki}ki_all.xlsx"
    write_harvest_all_by_month_sheets(inputs.harvest_template, all_path, df_ok, target_kis=[ki])
    print("[INFO] saved harvest all(UI, by month sheets):", all_path)

    # report
    df_report_unmapped = write_report_template(
        report_template=inputs.report_template,
        out_path=report_out,
        df_ok=df_ok,
        target_months=target_months,
        target_kis=target_kis,
        report_unit=report_unit,
        debug=debug,
    )
    if df_report_unmapped is not None and not df_report_unmapped.empty:
        df_ng = pd.concat([df_ng, df_report_unmapped], ignore_index=True)

    # unmapped
    write_unmapped(unmapped_out, df_ng)

    print("[INFO] saved report:", report_out)
    print("[INFO] saved unmapped:", unmapped_out)


# ============================================================
# CLI
# ============================================================
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser()

    p.add_argument("--norm-src", required=True)
    p.add_argument("--master", required=True)
    p.add_argument("--harvest-template", required=True)
    p.add_argument("--report-template", required=True)
    p.add_argument("--ki", required=True, type=int)      # 基準期（ファイル名用/harvest_all用）
    p.add_argument("--kis", default=None)                # "58,59"（reportへ書き込む対象期）
    p.add_argument("--outdir", default="out")
    p.add_argument("--validate", action="store_true")
    p.add_argument("--report-unit", choices=["g", "kg"], default="kg")
    p.add_argument("--debug", action="store_true")

    mx = p.add_mutually_exclusive_group(required=True)
    mx.add_argument("--month", type=int)
    mx.add_argument("--months")

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    debug = args.debug

    inputs = Inputs(
        norm_src=resolve_path(args.norm_src, "norm_src"),
        master=resolve_path(args.master, "master"),
        harvest_template=resolve_path(args.harvest_template, "harvest_template"),
        report_template=resolve_path(args.report_template, "report_template"),
    )
    outdir = ensure_outdir(args.outdir)
    target_months = parse_target_months(args.month, args.months)
    target_kis = parse_kis(args.kis, args.ki)

    run_pipeline(
        inputs=inputs,
        outdir=outdir,
        ki=int(args.ki),
        target_months=target_months,
        target_kis=target_kis,
        do_validate=args.validate,
        report_unit=args.report_unit,
        debug=debug,
    )


if __name__ == "__main__":
    main()
