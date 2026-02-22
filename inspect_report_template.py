from openpyxl import load_workbook

wb = load_workbook("report_template.xlsx", data_only=False)

print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("\n===", name, "===")
    # 「10月〜9月」が並ぶ行を探す（最大200行×50列）
    months = ["10月","11月","12月","1月","2月","3月","4月","5月","6月","7月","8月","9月"]
    hits = []
    for r in range(1, 201):
        row_vals = []
        for c in range(1, 51):
            v = ws.cell(r,c).value
            if isinstance(v,str):
                v = v.strip()
            row_vals.append(v)
        # 10月があって9月がある行を候補にする
        if "10月" in row_vals and "9月" in row_vals:
            hits.append((r, row_vals.index("10月")+1, row_vals.index("9月")+1))
    print("month header candidates:", hits[:5])

    # 「58期」「59期」のセルをざっくり探す（200行×30列）
    found = []
    for r in range(1, 201):
        for c in range(1, 31):
            v = ws.cell(r,c).value
            if isinstance(v,str):
                t = v.strip()
                if t in ("58期","59期","増減","稼働棟数"):
                    found.append((t, r, c))
    print("labels (58/59 etc) sample:", found[:30])
